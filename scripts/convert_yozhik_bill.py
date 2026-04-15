from pathlib import Path
import json
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE_DIR / "Еж и Билл.xlsx"
YOZHIK_PATH = BASE_DIR / "yozhik.json"
BILL_PATH = BASE_DIR / "bill.json"

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Ежик
daily = {}
ws_daily = wb["Ежедневные размышления"]

for row in ws_daily.iter_rows(min_row=2, values_only=True):
    date_value, text_value = row
    if date_value and text_value:
        key = str(date_value).strip()
        value = str(text_value).strip()
        daily[key] = value

# Билл
bill = {}
ws_bill = wb["Как это видит Билл"]

for row in ws_bill.iter_rows(min_row=2, values_only=True):
    num_value, text_value = row
    if num_value and text_value:
        key = str(int(num_value)).strip()
        value = str(text_value).strip()
        bill[key] = value

YOZHIK_PATH.write_text(
    json.dumps(daily, ensure_ascii=False, indent=2),
    encoding="utf-8"
)

BILL_PATH.write_text(
    json.dumps(bill, ensure_ascii=False, indent=2),
    encoding="utf-8"
)

print(f"Готово: {len(daily)} записей в yozhik.json")
print(f"Готово: {len(bill)} записей в bill.json")
